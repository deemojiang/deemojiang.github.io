# -*- coding: utf-8 -*-
"""从上级目录「安全」文件夹中的隐患汇总表生成 anquan_data.js（供大屏 file:// 直接打开）。"""
from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import urlparse

HERE = Path(__file__).resolve().parent
SECURITY_DIR = HERE.parent / "安全"
OUT_JS = HERE / "anquan_data.js"
OUT_JSON = HERE / "anquan_from_xlsx.json"


def _str(v):
    if v is None:
        return ""
    return str(v).strip()


def _pick(d: dict, *keys: str) -> str:
    for k in keys:
        if k in d and _str(d[k]):
            return _str(d[k])
    return ""


def _parse_url(url: str) -> tuple[str, str]:
    s = _str(url)
    if not s:
        return "", ""
    if "://" not in s and "://" not in s.lower():
        return "", ""
    try:
        p = urlparse(s if "://" in s else "http://" + s)
        host = p.hostname or ""
        port = str(p.port) if p.port else ""
        return host, port
    except Exception:
        return "", ""


def row_to_record(header: list[str], row: tuple) -> dict:
    raw: dict[str, object] = {}
    for i, key in enumerate(header):
        k = _str(key) or f"_列{i + 1}"
        val = row[i] if i < len(row) else None
        if val is None:
            raw[k] = ""
        elif isinstance(val, float) and val == int(val):
            raw[k] = int(val)
        else:
            raw[k] = val

    ent = _pick(raw, "单位名称")
    time = _pick(raw, "隐患发现时间", "发生时间")
    level = _pick(raw, "隐患等级", "紧急程度")
    src = _pick(raw, "事件来源")
    rule_type = _pick(raw, "隐患类型")
    rule_name = _pick(raw, "隐患描述", "事件标题", "通报基本情况")
    action = _pick(raw, "事件状态")
    atk = _pick(raw, "出现数量") or "1"
    domain = _pick(raw, "网站域名", "隐患域名url")
    ip = _pick(raw, "ip")
    port = _pick(raw, "端口")
    vuln_no = _pick(raw, "漏洞编号")

    uh, up = _parse_url(domain)
    if not ip:
        ip = uh
    if not port:
        port = up

    dst_display = domain or ip or "—"

    td_raw = raw.get("通报反馈天数")
    td_s = _str(td_raw)
    if not td_s:
        alarm_iv = "—"
    elif isinstance(td_raw, (int, float)):
        alarm_iv = f"{int(td_raw)}天" if td_raw == int(td_raw) else f"{td_raw}天"
    elif td_s.replace(".", "", 1).isdigit():
        alarm_iv = f"{int(float(td_s))}天"
    else:
        alarm_iv = td_s

    return {
        "ent": ent or "—",
        "time": time or "—",
        "level": level or "—",
        "srcIp": ip or "—",
        "srcPort": port or "—",
        "dstIp": dst_display,
        "dstPort": "—",
        "proto": "TCP",
        "policyId": vuln_no or "—",
        "ruleName": rule_name or "—",
        "ruleType": rule_type or "其他",
        "atkCount": atk,
        "hitCfg": _pick(raw, "通报等级") or "—",
        "alarmIv": alarm_iv,
        "action": action or "—",
        "src": src or "—",
        "_raw": raw,
    }


def main():
    candidates = sorted(SECURITY_DIR.glob("*隐患汇总*.xlsx"))
    if not candidates:
        raise SystemExit(f"未找到隐患汇总 xlsx：{SECURITY_DIR}")

    path = candidates[0]
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise SystemExit("工作表为空")

    header = [_str(c) or f"_列{i + 1}" for i, c in enumerate(rows[0])]
    data_rows = rows[1:]
    records = []
    for r in data_rows:
        if r is None or all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        records.append(row_to_record(header, list(r)))

    OUT_JSON.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")

    payload = json.dumps(records, ensure_ascii=False)
    payload = payload.replace("</script", "<\\/script")
    js = "/* 自动由 _export_anquan_data.py 自「安全」隐患汇总表生成，请勿手改 */\n"
    js += f"window.__EQS_ANQUAN__ = {payload};\n"
    OUT_JS.write_text(js, encoding="utf-8")

    print("源表:", path.name, "记录数:", len(records))
    print("已写:", OUT_JS.name, OUT_JSON.name)


if __name__ == "__main__":
    main()
