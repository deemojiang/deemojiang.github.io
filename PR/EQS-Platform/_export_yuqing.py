"""Export 长兴天能舆情信息.xlsx to mockYuqing JSON for EQS-Platform."""
from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path

import openpyxl


def norm_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    s = str(v).strip()
    return s


def sentiment_to_level(s: str) -> str:
    if s == "负面":
        return "高"
    if s == "中性":
        return "中"
    if s == "正面":
        return "低"
    return "中"


def sentiment_to_status(s: str) -> str:
    if s == "负面":
        return "待研判"
    if s == "正面":
        return "已监测"
    return "监测中"


def main() -> None:
    root = Path(__file__).resolve().parent.parent
    xlsx = list(root.glob("*舆情*.xlsx"))
    xlsx = [p for p in xlsx if not p.name.startswith("~")]
    if not xlsx:
        raise SystemExit("No *舆情*.xlsx under " + str(root))
    path = xlsx[0]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [c.value for c in ws[1]]
    out = []
    sent_c = Counter()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        d = {hdr[i]: r[i] for i in range(len(hdr)) if hdr[i]}
        sid = d.get("序号")
        try:
            rid = int(sid) if sid is not None else len(out) + 1
        except (TypeError, ValueError):
            rid = len(out) + 1
        title = norm_str(d.get("标题"))
        t = d.get("发布时间")
        if hasattr(t, "strftime"):
            time_s = t.strftime("%Y-%m-%d %H:%M:%S")
        else:
            time_s = norm_str(t)
        qx = norm_str(d.get("倾向性"))
        sent_c[qx] += 1
        site = norm_str(d.get("来源网站")) or "—"
        summary_raw = norm_str(d.get("摘要")) or title
        summary = summary_raw[:200] + ("…" if len(summary_raw) > 200 else "")
        kw = norm_str(d.get("关键词")) or "长兴天能"
        ent = kw if kw else "长兴天能"
        media = norm_str(d.get("媒体类型")) or "其他"
        row = {
            "id": rid,
            "time": time_s,
            "title": title,
            "ent": ent,
            "summary": summary,
            "sentiment": qx,
            "level": sentiment_to_level(qx),
            "src": site,
            "status": sentiment_to_status(qx),
            "channel": media,
            "author": norm_str(d.get("作者")),
            "region": norm_str(d.get("信息发布地域")),
            "articleType": norm_str(d.get("文章类型")),
            "publisherType": norm_str(d.get("发布者性质")),
            "reads": norm_str(d.get("阅读数")),
            "comments": norm_str(d.get("评论/回复数")),
            "reposts": norm_str(d.get("转发数")),
            "likes": norm_str(d.get("点赞数")),
            "fans": norm_str(d.get("粉丝量")),
            "url": norm_str(d.get("URL")),
            "keywords": kw,
        }
        out.append(row)
    wb.close()

    meta = {"source_file": path.name, "count": len(out), "sentiment": dict(sent_c)}
    dump = {"meta": meta, "mockYuqing": out}
    out_json = Path(__file__).resolve().parent / "yuqing_from_xlsx.json"
    out_json.write_text(json.dumps(dump, ensure_ascii=False, indent=2), encoding="utf-8")
    print("Wrote", out_json, "rows", len(out), "sentiment", dict(sent_c))


if __name__ == "__main__":
    main()
